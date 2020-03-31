#!/usr/bin/env python3
import time
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

courses = Workbook()
st_sheet = {}   # 用于导入students表中的数据，以课程名称为key
time_sheet = {}  # 用于导入time表中的数据，也以课程名称为key
out_data = {} # 整合后的字典，存储combine中待写数据

def combine():
	courses = load_workbook('courses.xlsx')

	# 提取students表中的数据
	students = courses['students']
	# 没有中文输入，所以将表头存入变量
	ct_title = students['A1'].value
	kn_title = students['B1'].value
	sn_title = students['C1'].value
#	print(courses.sheetnames)
	for row in students.rows:
		st_sheet[row[1].value] = [row[0].value, row[2].value]     # row[2]是创建时间
	del st_sheet[kn_title]     # 删除存有表头信息的键值对
#	print(st_sheet)

    	# 提取time表中的数据
	time = courses['time']
	st_title = time['C1'].value    # 存储学习时间的表头中文信息
	for row in time.rows:
		time_sheet[row[1].value] = [row[2].value]
	del time_sheet[kn_title]   # 删除存有表头信息的键值对
#	print(time_sheet)
	
	# 合并前两个字典后生成待导入的数据字典
	for k, n in st_sheet.items():
		for kn, t in time_sheet.items():
			if k == kn:
				out_data[k] = [n[0], n[1], t[0]]    # n[0]是创建时间,类型为datetime

	# 建立combine表,输出字典中的数据到新表
	combine = courses.create_sheet('combine')
	# 写入表头信息
	combine['A1'] = ct_title  # 创建时间
	combine['B1'] = kn_title  # 课程名称
	combine['C1'] = sn_title  # 学生人数
	combine['D1'] = st_title  # 学习时间

	# 将字典数据写入combine表中
	for k, l in  out_data.items():
		l.insert(1, k)
		l[0] = l[0].strftime('%Y-%m-%d %H:%M')    #将datatime.datatime类型转为str类型，否则有小数出现
		combine.append(l)

	# 写入Excel文件
	courses.save('courses.xlsx')
	courses.close()
'''
	# 由于不能查看Excel文件，所以测试表中数据是否准确
	print('sheetnames: ', courses.sheetnames)
	cell = combine['A1':'D11']
	for item in cell:
		for ce in item:
			print(ce.value)
'''

"""
	用于拆分年份后生成工作簿的函数
"""

def split():
	courses = load_workbook('courses.xlsx')
	combine = courses['combine']
	comb_dict = {}



	#提取表头信息
	ct_title = combine['A1'].value
	kn_title = combine['B1'].value
	sn_title = combine['C1'].value
	st_title = combine['D1'].value

	# 建立一个键为“创建时间“的字典，用于区分年份
	for each in combine.rows:
		comb_dict[each[0].value] = [each[1].value, each[2].value, each[3].value]   #each[0]的时间为str类型
	del comb_dict[ct_title]   # 删除表头信息键值对
	courses.close()

	year = list(comb_dict.keys())   #将字典中的键转为列表，否则遍历时操作字典会报错
	# 提取一个只包含年份的列表，用于判断字典键列表的年份
	year = [datetime.datetime.strptime(y, '%Y-%m-%d %H:%M').strftime('%Y') for y in year]  #解析式提取纯年份列表
	year = list(set(year))   # 列表去重


	for title in year:
		new_wb_name = str(title) + '.xlsx'    #生成一个文件名 年份+ 扩展名
		new_wb = Workbook()                   #实例化新工作簿
		new_ws = new_wb.create_sheet(str(title), 0)     #用年份名字新建一个表
		new_ws['A1'] = ct_title    #写入表头
		new_ws['B1'] = kn_title
		new_ws['C1'] = sn_title
		new_ws['D1'] = st_title
		new_wb.remove(new_wb['Sheet'])     #删除新建工作表时的sheet默认表


		# 用纯年份列表year和字典键列表进行比较，查找符合相同年份的列表值：y
		for y in list(comb_dict.keys()):
			if title == datetime.datetime.strptime(y, '%Y-%m-%d %H:%M').strftime('%Y'):  # 转换类型为纯年份后与year比较
				output_list = comb_dict.get(y)  # 为什么一定要把字典返回的get值赋给一个半变量，才能再用insert
				output_list.insert(0, y)  # 如果用comb_dict.get(y).insert(0,y)始终返回None ?,此时.get(y)明明有值
				new_ws.append(output_list)  # 添加数据到工作表
			new_wb.save(new_wb_name)      #存储到新的按年分割的文件中
		del comb_dict[y]  # 在数据总字典中删除已匹配年份的键值对




if __name__ == '__main__':
	combine()
	split()
	 
